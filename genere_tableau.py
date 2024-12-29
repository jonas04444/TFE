import sqlite3
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

def export_comparaison_excel(version_actuelle, version_proposition, ligne, sens, fichier_sortie):
    try:

        connect = sqlite3.connect("listelieux.db")
        cursor = connect.cursor()

        colonnes = ["Lieu de Départ", "Lieu d'Arrivée", "Description Arrivée", "Distance (mètres)", "Type",
                    "0:00", "7:00", "9:00", "15:30", "17:30"]

        requeteSQL = """
        SELECT 
            pl.LieuxDepart AS "Lieu de Départ",
            pl.LieuxArrivee AS "Lieu d'Arrivée",
            nl.Description AS "Description Arrivée",
            pl.Distance AS "Distance (mètres)",
            tl0.Temps AS "0:00",
            tl7.Temps AS "7:00",
            tl9.Temps AS "9:00",
            tl15.Temps AS "15:30",
            tl17.Temps AS "17:30"
        FROM PaireLieux AS pl
        LEFT JOIN NomLieux AS nl ON pl.LieuxArrivee = nl.NomLieux
        LEFT JOIN TempsEntreLieux2 AS tl0 ON pl.IDPaireLieux = tl0.PaireLieux AND tl0.VersionTemps = ? AND tl0.HeureDebut = '0:00'
        LEFT JOIN TempsEntreLieux2 AS tl7 ON pl.IDPaireLieux = tl7.PaireLieux AND tl7.VersionTemps = ? AND tl7.HeureDebut = '7:00'
        LEFT JOIN TempsEntreLieux2 AS tl9 ON pl.IDPaireLieux = tl9.PaireLieux AND tl9.VersionTemps = ? AND tl9.HeureDebut = '9:00'
        LEFT JOIN TempsEntreLieux2 AS tl15 ON pl.IDPaireLieux = tl15.PaireLieux AND tl15.VersionTemps = ? AND tl15.HeureDebut = '15:30'
        LEFT JOIN TempsEntreLieux2 AS tl17 ON pl.IDPaireLieux = tl17.PaireLieux AND tl17.VersionTemps = ? AND tl17.HeureDebut = '17:30'
        WHERE pl.IDPaireLieux IN (
            SELECT c.IdPaireLieux 
            FROM composition AS c
            JOIN Ligne AS l ON c.IdLigne = l.IDLigne
            WHERE l.NumLigne = ? AND l.Sens = ?
        )
        ORDER BY pl.IDPaireLieux;
        """

        cursor.execute(requeteSQL, (version_actuelle,) * 5 + (ligne, sens))
        donnees_actuelles = cursor.fetchall()

        cursor.execute(requeteSQL, (version_proposition,) * 5 + (ligne, sens))
        donnees_proposition = cursor.fetchall()

        tableau = []
        for actuel, proposition in zip(donnees_actuelles, donnees_proposition):
            lieu_depart = actuel[0]
            lieu_arrivee = actuel[1]
            description = actuel[2]
            distance = actuel[3]

            tableau.append([lieu_depart, lieu_arrivee, description, distance, "Temps Actuel"] + list(actuel[4:]))
            tableau.append([lieu_depart, lieu_arrivee, description, distance, "Vitesse Actuelle"] +
                           [round((float(distance) / 1000) / (float(t) / 60), 2) if t and t != '' else "" for t in actuel[4:]])
            tableau.append([lieu_depart, lieu_arrivee, description, distance, "Temps Proposition"] + list(proposition[4:]))
            tableau.append([lieu_depart, lieu_arrivee, description, distance, "Vitesse Proposition"] +
                           [round((float(distance) / 1000) / (float(t) / 60), 2) if t and t != '' else "" for t in proposition[4:]])

        df = pd.DataFrame(tableau, columns=colonnes)

        wb = Workbook()
        ws = wb.active
        ws.title = "Comparaison"

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            ws.append(row)

        grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        thick_border = Border(
            left=Side(style="thick"),
            right=Side(style="thick"),
            top=Side(style="thick"),
            bottom=Side(style="thick")
        )

        current_lieu = None
        current_start = None

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(tableau) + 1, min_col=1, max_col=4), start=2):
            lieu_cell = row[0]
            if lieu_cell.value != current_lieu:
                if current_lieu is not None:
                    ws.merge_cells(start_row=current_start, start_column=1, end_row=row_idx - 1, end_column=1)
                    ws.merge_cells(start_row=current_start, start_column=2, end_row=row_idx - 1, end_column=2)
                    ws.merge_cells(start_row=current_start, start_column=3, end_row=row_idx - 1, end_column=3)
                    ws.merge_cells(start_row=current_start, start_column=4, end_row=row_idx - 1, end_column=4)
                current_lieu = lieu_cell.value
                current_start = row_idx

        if current_lieu is not None:
            ws.merge_cells(start_row=current_start, start_column=1, end_row=len(tableau) + 1, end_column=1)
            ws.merge_cells(start_row=current_start, start_column=2, end_row=len(tableau) + 1, end_column=2)
            ws.merge_cells(start_row=current_start, start_column=3, end_row=len(tableau) + 1, end_column=3)
            ws.merge_cells(start_row=current_start, start_column=4, end_row=len(tableau) + 1, end_column=4)

        for row in ws.iter_rows(min_row=2, max_row=len(tableau) + 1):
            type_cell = row[4]
            if type_cell.value == "Temps Proposition":
                for col_idx, cell in enumerate(row[5:], start=6):
                    correspondant = ws.cell(row=cell.row - 2, column=col_idx)
                    if cell.value and correspondant.value and isinstance(cell.value, (int, float)) and isinstance(
                            correspondant.value, (int, float)):
                        if float(cell.value) < float(correspondant.value):
                            cell.fill = red_fill
                        elif float(cell.value) > float(correspondant.value):
                            cell.fill = green_fill
                        else:
                            cell.fill = white_fill
            elif type_cell.value == "Temps Actuel":
                for cell in row[5:]:
                    cell.fill = grey_fill

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(fichier_sortie)
        print(f"Fichier exporté avec succès : {fichier_sortie}")

    except sqlite3.Error as e:
        print(f"Erreur SQL : {e}")
    except Exception as ex:
        print(f"Erreur inattendue : {ex}")
    finally:
        if connect:
            connect.close()

export_comparaison_excel("202412", "202501", 41, 2, "comparaison_lignes.xlsx")
